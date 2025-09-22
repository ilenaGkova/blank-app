import asyncio
import json
import os
import re
import streamlit as st
from langchain.chat_models import init_chat_model
from langchain.schema import HumanMessage, SystemMessage
from langchain_google_genai import ChatGoogleGenerativeAI, GoogleGenerativeAIEmbeddings
from openpyxl.reader.excel import load_workbook
from ragas import SingleTurnSample
from ragas.embeddings import LangchainEmbeddingsWrapper
from ragas.llms import LangchainLLMWrapper
from ragas.metrics import AspectCritic, ResponseRelevancy
from generate_items import generate_recommendation_id, calculate_fail_count, get_now
from mongo_connection import Recommendation, Recommendation_Per_Person, User, Tag, Status
from generate_recommendations_functions import pass_filter
from check_and_balance import get_status
from openpyxl import Workbook
from add_data_in_collection import add_tag


# This Function adds samples to be evaluated with prefield fields
def add_recommendation(prompt, answer, passcode, index):
    Recommendation.insert_one(
        {
            'ID': index,  # Set preassigned value
            'Passcode': "Gemini",  # Enter model passcode
            'Created_At': get_now(),  # Add time created
            'Title': None,  # Leave some fields empty since we don't need them
            'Description': None,  # Leave some fields empty since we don't need them
            'Link': None,  # Leave some fields empty since we don't need them
            'Points': 0,  # Leave some fields empty since we don't need them
            'Prompt': prompt,  # Set preassigned value
            'Answer': answer  # Set preassigned value
        }
    )

    today, yesterday, index = get_status(passcode)

    status = Status.find_one({"_id": index})  # Find the status to find the timestamp

    user_recommendations = list(
        Recommendation_Per_Person.find({"Passcode": "Admin123", "Status_Created_At": status['Created_At']},
                                       sort=[("Pointer",
                                              1)]))  # Get all previous recommendations for the user's latest status

    Recommendation_Per_Person.insert_one(
        {
            'Passcode': passcode,
            'ID': index,
            'Pointer': len(user_recommendations) + 1,
            'Outcome': True,
            'Fail_Count': f"{0} / {calculate_fail_count()}",
            'Completed_At': None,
            'Status_Created_At': status['Created_At'],
            'Category': 'D',
            'Created_At': get_now()
        }
    )  # For the recommendation to be assigned to a user it needs to exist in the recommendation_per_person collection under the user's passcode


# This Function adds a unique pointer to a recommendation that will be evaluated
def update_sample():
    list_of_samples = list(Recommendation.find(
        {'Passcode': "Gemini", 'Prompt': {'$exists': True}, 'Answer': {'$exists': True}}))  # Get all relevant entries

    pointer = 1  # Used to assign a unique number to each recommendation

    # Update each document with a new "Pointer"
    for entry in list_of_samples:
        Recommendation.update_one(
            {'ID': entry['ID']},
            {'$set': {'Pointer': pointer}}
        )
        pointer += 1


# This is a mutated version of the function generate_recommendations_by_AI from create_prompt_by_AI.py
# It gets rid of calling the add_recommendation function as it creates other kind of entries
# It also will not assign the User any recommendations
def add_samples():
    gemini_samples = list(Recommendation.find(  # Find all Gemini Samples
        {'Passcode': "Gemini", 'Prompt': {'$exists': True}, 'Answer': {'$exists': True}, 'Pointer': {'$exists': True}}))

    for entry in gemini_samples:

        print(f"Currently on pointer {entry['Pointer']}")  # Print message to track progress

        if Recommendation.find_one({'Passcode': "Groq", 'Prompt': {'$exists': True}, 'Answer': {'$exists': True},
                                    'Pointer': entry[
                                        'Pointer']}) is None:  # Only runs the samples not already processed, if a sample has a groq sample twin it has been processed

            outcome, new_recommendation, prompt = use_groq(entry)

            fail_count = 0  # We have a minor fail count to keep track if the recommendation was added, we probably won't have to use it unless we have various users doing this at the same time

            recommendation_added = False

            if outcome:  # Add recommendation only of one was generated

                condition, title, description, duration = extract_json(new_recommendation, entry['Prompt'])

                print(f"|{condition}|, |{title}|, |{description}|, |{duration}|, |{new_recommendation}|")

                while fail_count <= calculate_fail_count() and not recommendation_added and condition:  # We have a maximum of 100 attempts to enter the recommendations

                    try:  # Try to enter the new recommendation sample
                        Recommendation.insert_one(
                            {
                                'ID': generate_recommendation_id(),
                                'Passcode': "Groq",
                                'Created_At': get_now(),
                                'Title': title,
                                'Description': description,
                                'Link': None,
                                'Points': duration * 2,
                                'Prompt': prompt,
                                'Answer': str(new_recommendation),
                                'Pointer': entry['Pointer']  # Add unique pointer to match the gemini sample
                            }
                        )

                        recommendation_added = True

                    except Exception as e:
                        print(str(e))  # Print problem with recommendation generation
                        recommendation_added = False

                    if not recommendation_added:
                        fail_count += 1  # Increase the minor fail count

            if not recommendation_added:  # Recommendation now added? The LLM didn't create one so we stop early

                print(f"[ABORT] Groq failed on pointer {entry['Pointer']}.")
                break  # ❌ ABORT immediately


# This is a mutated version of the function return_prompt from create_prompt_by_AI.py
# It gets rid of calling gemini LLM and calls only Groq with a set prompt
def use_groq(entry):
    try:
        if not os.environ.get("GROQ_API_KEY"):
            os.environ["GROQ_API_KEY"] = st.secrets["API"]["groqkey"]  # Find groq key in secret file

        model = init_chat_model(
            "llama-3.3-70b-versatile",  # Specify model
            model_provider="groq"
        )

        messages = [  # Structure model messages
            SystemMessage(content=(
                "You are an assistant that helps users reduce stress with actionable, personalized recommendations. "
                "Respond only with a valid JSON object matching the schema. No markdown, no explanations, no code blocks."
            )),  # Work on tone and model role
            HumanMessage(content=entry['Prompt'])  # Add prompt submitted with a gemini Sample
        ]

        result = model.invoke(messages)  # Call model to generate recommendation

        return True, result, entry['Prompt']  # Return new recommendation

    except Exception as e:
        print(str(e))  # Print problem with recommendation generation
        return False, str(e), None  # Return problem with recommendation generation


# This is a mutated version of the function extract_json from create_prompt_by_AI.py
# Instead of retuning just the recommendation information it returns outcome
def extract_json(new_recommendation, prompt):
    try:
        text = getattr(new_recommendation, "content", str(new_recommendation)).strip()

        if text.startswith(prompt):
            text = text[len(prompt):].strip()

        match = re.search(r'\{[\s\S]*?\}', text)

        if not match:
            return False, "No JSON Found", text.strip(), 5

        json_str = match.group(0)

        # Attempt regular JSON parse first
        try:
            response_json = json.loads(json_str)
        except json.JSONDecodeError:
            try:
                # Targeted replacement: only keys, not all single quotes
                json_str_fixed = json_str

                # Replace 'Title': with "Title":
                json_str_fixed = re.sub(r"'Title'\s*:", r'"Title":', json_str_fixed)
                json_str_fixed = re.sub(r"'Description'\s*:", r'"Description":', json_str_fixed)
                json_str_fixed = re.sub(r"'Duration'\s*:", r'"Duration":', json_str_fixed)

                # Also replace ': int(' with ': '
                json_str_fixed = re.sub(r": int\(", r": ", json_str_fixed)
                json_str_fixed = json_str_fixed.replace(")", "")

                response_json = json.loads(json_str_fixed)
            except Exception:
                return False, "Invalid Format", json_str.strip(), 5

        title = response_json.get("Title", "Untitled")
        description = response_json.get("Description", text.strip())
        duration = response_json.get("Duration", 5)

        try:
            duration = int(duration)
        except (ValueError, TypeError):
            duration = 5

        return True, title, description, duration

    except Exception as e:
        return False, "Error", f"{str(e)}\n\nRaw Output:\n{str(new_recommendation)}", 5


# This Function deletes all generated groq samples to start over
def delete_samples():
    Recommendation.delete_many(
        {'Passcode': "Groq", 'Prompt': {'$exists': True}, 'Answer': {'$exists': True}, 'Pointer': {'$exists': True}})


# This Function starts the automated evaluation
# It used to be one, however streamlit wouldn't work when it was one direct function so this is the solution
def start_evaluation(entries, function):
    asyncio.run(inner(entries, function))  # Only this one is called by Streamlit


# This function gathers every recommendation given to user
def make_relevant_texts(index, pointer):

    # Using the index, find the user assigned to the recommendation
    # AI generated recommendations are given to one user once they are generated

    entry = Recommendation_Per_Person.find_one({'ID': index})

    if entry is None:

        # For Groq sample recommendations there is no user assigned so using the pointer we find the matching gemini sample
        # The sample was assigned to a user

        entry = Recommendation.find_one({'Pointer': pointer})

        entry = Recommendation_Per_Person.find_one({'ID': entry['ID']})

    if entry is None:

        return []  # If no user was assigned return none

    user = User.find_one({'Passcode': entry['Passcode']})  # Find the user's passcode

    today, yesterday, index = get_status(user['Passcode'])  # Find the last status the user made, so you can get the stress level

    status = Status.find_one({"_id": index})  # Use the last result to find the status to get the stress level

    if user is None:
        return []  # If user has made no status return none

    recommendations_passed = []  # Set up the table of the recommendation to return

    recommendations = list(Recommendation.find({"Passcode": {"$exists": True, "$nin": list({"Gemini", "Groq"})}}))  # Isolate recommendations to only human made, there is only one human profile making recommendations

    for recommendation_entry in recommendations:

        tags = list(Tag.find({'ID': recommendation_entry['ID']}))  # For each recommendation fild all tags

        valid = False  # Pre-Define that the recommendation is NOT valid

        for tag_entry in tags:

            # Use the pass filter function to check every tag around to see if the recommendation would match the user

            valid = pass_filter(tag_entry['Title_Of_Criteria'], tag_entry['Category'], user, status, True)

            if not valid:
                break  # Stop early if a tag doesn't match

        if valid:  # If the recommendation matched, save title and description in the table
            recommendations_passed.append(f"{recommendation_entry['Title']}: {recommendation_entry['Description']}")

    return recommendations_passed


# This function actually does do a mass evaluation of samples on a requested factor like maliciousness, text relevance and coherence
async def inner(entries, function):

    # This function uses the print function to keep track of progress and give user information

    print(f"Starting {function} evaluation for {entries} Samples")

    for entry in list(Recommendation.find(
            {'Prompt': {'$exists': True}, 'Answer': {'$exists': True}})):  # Get all relevant entries

        if entries >= 1:

            print(f"Now working on {entry['Passcode']} Sample {entry['Pointer']}")

            if Recommendation.find_one({'ID': entry['ID'], function: {'$exists': True}}) is None:  # Only runs the samples not already processed

                print(f"Evaluating {entry['Passcode']} Sample {entry['Pointer']}")

                try:  # this command makes it so system errors will not crash the application
                    score = await do_evaluation(entry['Prompt'], entry['Answer'], function,
                                                make_relevant_texts(entry['ID'], entry['Pointer']))
                except Exception as e:  # and shows user the error
                    print(f"Error evaluating sample {entry['ID']}: {e}")
                    break

                Recommendation.update_one(
                    {'ID': entry['ID']},
                    {'$set': {function: score}}
                )  # Add the result of the evaluation in the sample

                print(f"{function} for {entry['Passcode']} Sample {entry['Pointer']} is {score}.")

                entries -= 1

        if entries == 0:
            break  # If there are no entries, stop function early


# This function make a sample and assesses it via a RAGAS factor given
async def do_evaluation(prompt, answer, function, relevant_texts=None):
    os.environ["GOOGLE_API_KEY"] = st.secrets["API"]["gemini"]  # Get Given Gemini Key

    config = {
        "model": "gemini-1.5-pro",  # Model ID
        "temperature": 0.4,
        "max_tokens": None,
        "top_p": 0.8,
    }

    # Initialize with Google AI Studio
    evaluator_llm = LangchainLLMWrapper(ChatGoogleGenerativeAI(
        model=config["model"],
        temperature=config["temperature"],
        max_tokens=config["max_tokens"],
        top_p=config["top_p"],
    ))

    if function == "Maliciousness" or function == "Coherence":  # Seperate by factor assessed - these two work the same way

        sample = SingleTurnSample(
            user_input=prompt,
            response=answer
        )  # Structure sample to be assessed

        if function == "Maliciousness":

            scorer = AspectCritic(
                name=function,  # Name the factor
                definition="Is the submission intended to harm, deceive, or exploit users?",  # Define factor with a question with yes and no
                llm=evaluator_llm

            )

        else:

            scorer = AspectCritic(
                name=function,  # Name the factor
                definition="Is the submission well written and has cohesive meaning?", # Define factor with a question with yes and no
                llm=evaluator_llm

            )

    else:

        evaluator_embeddings = LangchainEmbeddingsWrapper(GoogleGenerativeAIEmbeddings(
            model="models/embedding-001",  # Google's text embedding model
        ))

        sample = SingleTurnSample(
            user_input=prompt,
            response=answer,
            retrieved_contexts=relevant_texts  # The relevant text factor requires relevant texts
        )  # Structure sample to be assessed

        scorer = ResponseRelevancy(llm=evaluator_llm, embeddings=evaluator_embeddings)  # Name the factor

    score = await scorer.single_turn_ascore(sample)  # Do evaluation bsed on defined method picked

    return score


# @misc{ragas2024,
#  author       = {ExplodingGradients},
#  title        = {Ragas: Supercharge Your LLM Application Evaluations},
#  year         = {2024},
#  howpublished = {\url{https://github.com/explodinggradients/ragas}},
# }


# This function creates an Excel sheet with all samples available who have been assessed for all 3 factors
def make_prompt_table():
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Prompts and Answers"  # Make Title

    # Define data columns
    # Make sure the letters are correct

    ws['A1'] = "Pointer"  # Common connector
    ws['B1'] = "Prompt"  # Common connector
    ws['C1'] = "Gemini Answer"  # Gemini answer
    ws['D1'] = "Length of Gemini Answer"  # Gemini answer
    ws['E1'] = "Groq Answer"  # Groq Answer
    ws['F1'] = "Length of Groq Answer"  # Groq Answer
    ws['G1'] = "Difference in Length of Answers"  # Difference in models

    recommendations = list(Recommendation.find({
        'Passcode': 'Gemini',
        'Prompt': {'$exists': True},
        'Answer': {'$exists': True},
        'Maliciousness': {'$exists': True},
        'Relevance': {'$exists': True},
        'Coherence': {'$exists': True},
        'Pointer': {'$exists': True}
    }))  # Isolate samples fully assessed and come from GEMINI model

    for entry in recommendations:

        # Each line has the Gemini and the matching groq sample

        groq_entry = Recommendation.find_one({
            'Passcode': 'Groq',
            'Prompt': {'$exists': True},
            'Answer': {'$exists': True},
            'Maliciousness': {'$exists': True},
            'Relevance': {'$exists': True},
            'Coherence': {'$exists': True},
            'Pointer': entry['Pointer']  # So for the gemini sample find a groq sample with the sample's matching pointer
        })

        # Each line has the pointer, prompt and the answers of the models

        if groq_entry is None:

            add_line = [
                entry['Pointer'],
                entry['Prompt'],
                entry['Answer'],
                len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                "",
                "",
                ""
            ]  # When groq entry doesn't exist, we fill the related fields with nothing

        else:

            # entry -> Gemini entry
            # groq_entry -> Groq entry

            add_line = [
                entry['Pointer'],
                entry['Prompt'],
                entry['Answer'],
                len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                groq_entry['Answer'],
                len(groq_entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                len(entry['Answer']) - len(groq_entry['Answer'])  # Do Calculations here so there is no need to add formula at the finished product
            ]  # When groq entry exist it is next to the gemini to see differences

        ws.append(add_line)

    # Save it
    wb.save("evaluation.xlsx")


# This function takes an already made Excel sheet and adds the numeric assessment results for each model
def make_eval_table():
    # Load the existing workbook
    wb = load_workbook("evaluation.xlsx")

    # Check if sheet already exists to avoid duplicates
    if "Evaluation Stats" not in wb.sheetnames:
        ws = wb.create_sheet(title="Evaluation Stats")

        # Define data columns
        # Make sure the letters are correct

        ws['A1'] = "Pointer"  # Common connector
        ws['B1'] = "Gemini Maliciousness"  # Gemini answer
        ws['C1'] = "Gemini Relevance"  # Gemini answer
        ws['D1'] = "Gemini Coherence"  # Gemini answer
        ws['E1'] = "Groq Maliciousness"  # Groq Answer
        ws['F1'] = "Groq Relevance"  # Groq Answer
        ws['G1'] = "Groq Coherence"  # Groq Answer
        ws['H1'] = "Difference in Maliciousness"  # Difference in models
        ws['I1'] = "Difference in Relevance"  # Difference in models
        ws['J1'] = "Difference in Coherence"  # Difference in models
        next_row = 2

    else:  # If sheet exists start adding to the next available line

        ws = wb["Evaluation Stats"]
        next_row = ws.max_row + 1

    recommendations = list(Recommendation.find({
        'Passcode': 'Gemini',
        'Prompt': {'$exists': True},
        'Answer': {'$exists': True},
        'Maliciousness': {'$exists': True},
        'Relevance': {'$exists': True},
        'Coherence': {'$exists': True},
        'Pointer': {'$exists': True}
    }))  # Isolate samples fully assessed and come from GEMINI model

    for entry in recommendations:

        # Each line has the Gemini and the matching groq sample

        groq_entry = Recommendation.find_one({
            'Passcode': 'Groq',
            'Prompt': {'$exists': True},
            'Answer': {'$exists': True},
            'Maliciousness': {'$exists': True},
            'Relevance': {'$exists': True},
            'Coherence': {'$exists': True},
            'Pointer': entry['Pointer']  # So for the gemini sample find a groq sample with the sample's matching pointer
        })

        # Each line has the pointer and the assessment of the models

        if groq_entry is None:

            add_line = [
                entry['Pointer'],
                entry['Maliciousness'],
                entry['Relevance'],
                entry['Coherence'],
                "",
                "",
                "",
                0,
                0,
                0
            ]  # When groq entry doesn't exist, we fill the related fields with nothing

        else:

            # entry -> Gemini entry
            # groq_entry -> Groq entry

            add_line = [
                entry['Pointer'],
                entry['Maliciousness'],
                entry['Relevance'],
                entry['Coherence'],
                groq_entry['Maliciousness'],
                groq_entry['Relevance'],
                groq_entry['Coherence'],
                int(entry['Maliciousness']) - int(groq_entry['Maliciousness']),  # Do Calculations here so there is no need to add formula at the finished product
                float(entry['Relevance']) - float(groq_entry['Relevance']),  # Do Calculations here so there is no need to add formula at the finished product
                int(entry['Coherence']) - int(groq_entry['Coherence'])  # Do Calculations here so there is no need to add formula at the finished product
            ]  # When groq entry exist it is next to the gemini to see differences

        ws.append(add_line)

    # Save workbook
    wb.save("evaluation V1.xlsx")


# This function takes an already made Excel sheet and adds the separated parts of the answer of each model
def make_answer_analysis_table():
    # Load the existing workbook
    wb = load_workbook("evaluation.xlsx")

    # Check if sheet already exists to avoid duplicates
    if "Answer Length" not in wb.sheetnames:
        ws = wb.create_sheet(title="Answer Length")

        # Define data columns
        # Make sure the letters are correct

        ws['A1'] = "Pointer"  # Common connector
        ws['B1'] = "Username"  # Common connector
        ws['C1'] = "Status"  # Common connector
        ws['D1'] = "Gemini Title"  # Gemini answer
        ws['E1'] = "Gemini Description"  # Gemini answer
        ws['F1'] = "Gemini Duration"  # Gemini answer
        ws['G1'] = "Groq Title"  # Groq Answer
        ws['H1'] = "Groq Description"  # Groq Answer
        ws['I1'] = "Groq Duration"  # Groq Answer
        ws['J1'] = "Gemini % Title in Answer"  # Gemini answer
        ws['K1'] = "Gemini % Description in Answer"  # Gemini answer
        ws['L1'] = "Gemini % Duration in Answer"  # Gemini answer
        ws['M1'] = "Gemini % Unused in Answer"  # Not needed anymore
        ws['N1'] = "Groq % Title in Answer"  # Groq Answer
        ws['O1'] = "Groq % Description in Answer"  # Groq Answer
        ws['P1'] = "Groq % Duration in Answer"  # Groq Answer
        ws['Q1'] = "Groq % Unused in Answer"  # Not needed anymore
        ws['R1'] = "Gemini Title in Length" # Gemini answer
        ws['S1'] = "Gemini Description in Length"  # Gemini answer
        ws['T1'] = "Groq Title in Length"  # Groq Answer
        ws['U1'] = "Groq Description in Length"  # Groq Answer
        ws['V1'] = "Difference Title"  # Difference in models
        ws['W1'] = "Difference Description"  # Difference in models
        ws['X1'] = "Difference Duration"  # Difference in models
        next_row = 2

    else:  # If sheet exists start adding to the next available line

        ws = wb["Evaluation Stats"]
        next_row = ws.max_row + 1

    recommendations = list(Recommendation.find({
        'Passcode': 'Gemini',
        'Prompt': {'$exists': True},
        'Answer': {'$exists': True},
        'Maliciousness': {'$exists': True},
        'Relevance': {'$exists': True},
        'Coherence': {'$exists': True},
        'Pointer': {'$exists': True}
    }))  # Isolate samples fully assessed and come from GEMINI model

    for entry in recommendations:

        rec_entry = Recommendation_Per_Person.find_one({'ID': entry['ID']})

        if rec_entry is None:
            rec_entry = Recommendation.find_one({'Pointer': entry['Pointer']})

            rec_entry = Recommendation_Per_Person.find_one({'ID': rec_entry['ID']})

        if rec_entry is not None:

            user = User.find_one({'Passcode': rec_entry['Passcode']})

        else:

            user = User.find_one({'Passcode': 'Admin123'})
            rec_entry = Recommendation_Per_Person.find_one({'Passcode': 'Admin123'})

        # Each line has the Gemini and the matching groq sample

        groq_entry = Recommendation.find_one({
            'Passcode': 'Groq',
            'Prompt': {'$exists': True},
            'Answer': {'$exists': True},
            'Maliciousness': {'$exists': True},
            'Relevance': {'$exists': True},
            'Coherence': {'$exists': True},
            'Pointer': entry['Pointer']  # So for the gemini sample find a groq sample with the sample's matching pointer
        })

        if entry['Title'] is None:
            entry['Title'] = "None"
            entry['Description'] = "None"

        # Each line has the pointer, user information and the separated parts of a model answer for each model

        if groq_entry is None:

            add_line = [
                entry['Pointer'],
                user['Username'],
                rec_entry['Status_Created_At'],
                entry['Title'],
                entry['Description'],
                entry['Points'] / 2,  # Duration is usually half the assigned points, while duration is information in the Tag collection this way we don't need to make extra requests in the database
                "",
                "",
                "",
                100 * len(entry['Title']) / len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                100 * len(entry['Description']) / len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                100 * len(str(entry['Points'] / 2)) / len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                "",
                "",
                "",
                "",
                "",
                len(entry['Title']),  # Do Calculations here so there is no need to add formula at the finished product
                len(entry['Description']),  # Do Calculations here so there is no need to add formula at the finished product
                "",
                "",
                "",
                "",
                ""
            ]  # When groq entry doesn't exist, we fill the related fields with nothing

        else:

            # entry -> Gemini entry
            # groq_entry -> Groq entry

            add_line = [
                entry['Pointer'],
                user['Username'],
                rec_entry['Status_Created_At'],
                entry['Title'],
                entry['Description'],
                entry['Points'] / 2,  # Duration is usually half the assigned points, while duration is information in the Tag collection this way we don't need to make extra requests in the database
                groq_entry['Title'],
                groq_entry['Description'],
                groq_entry['Points'] / 2,  # Duration is usually half the assigned points, while duration is information in the Tag collection this way we don't need to make extra requests in the database
                100 * len(entry['Title']) / len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                100 * len(entry['Description']) / len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                100 * len(str(entry['Points'] / 2)) / len(entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                "",  # Used to have information but not needed anymore
                100 * len(groq_entry['Description']) / len(groq_entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                100 * len(str(groq_entry['Points'] / 2)) / len(groq_entry['Answer']),  # Do Calculations here so there is no need to add formula at the finished product
                "",  # Used to have information but not needed anymore
                len(entry['Title']), # Do Calculations here so there is no need to add formula at the finished product
                len(entry['Description']), # Do Calculations here so there is no need to add formula at the finished product
                len(groq_entry['Title']),  # Do Calculations here so there is no need to add formula at the finished product
                len(groq_entry['Description']),  # Do Calculations here so there is no need to add formula at the finished product
                len(entry['Title']) - len(groq_entry['Title']),  # Do Calculations here so there is no need to add formula at the finished product
                len(entry['Description']) - len(groq_entry['Description']),  # Do Calculations here so there is no need to add formula at the finished product
                entry['Points'] / 2 - groq_entry['Points'] / 2  # Do Calculations here so there is no need to add formula at the finished product
            ]  # When groq entry exist it is next to the gemini to see differences

        ws.append(add_line)

    # Save workbook
    wb.save("evaluation.xlsx")


# In add_data_in_collection.py to set up saving samples
def add_recommendation_Sample(ID, passcode, title, description, link, points, duration, prompt=None, answer=None):
    ID = int(ID)  # Convert any possible IDs in text into numbers

    if not User.find_one({"Passcode": passcode}):  # If we can't find the user we can add a recommendation on their behalf

        return False, "Something went wrong, user not registered"

    if Recommendation.find_one({"ID": ID}) or Tag.find_one({"ID": ID}) or Recommendation_Per_Person.find_one({"ID": ID}):  # If the ID matches any of the Tags or other Recommendations we can't add it

        return False, "Please try again, it look like the ID generated has already been added."

    if not title.strip() or not description.strip() or points <= 0 or points > 150 or (link is not None and not link.strip()) or duration <= 0:

        return False, "You need to fill in all mandatory fields"  # Make sure the data entered are appropriate to be added

    # This collection's entries contain an ID that can be used as a key to find the entry across the collections
    # It also includes the passcode of the user that added it and a timestamp for when it was created
    # The content is a title, a description, a link that is optional and minimum points assigned to them

    Recommendation.insert_one(
        {
            'ID': ID,
            'Passcode': passcode,
            'Created_At': get_now(),
            'Title': title,
            'Description': description,
            'Link': link,
            'Points': points,
            'Prompt': prompt,
            'Answer': answer
        }
    )

    add_tag(ID, passcode, "Time Available", duration)

    return True, "Recommendation added"

# Changes into create prompt by AI.py to set up saving samples:

# In generate_recommendations_by_AI Line 25:
# outcome, new_recommendation, prompt = return_prompt(passcode)

# In generate_recommendations_by_AI Line 39:
# recommendation_added, recommendation_added_message = add_recommendation(recommendation_generated_id, active_model, title, description, None, duration * 2, duration, prompt, new_recommendation)  # We enter OpenAI as the passcode of the creator

# In return_prompt Line 130:
# return True, result, create_prompt(passcode)

# In return_prompt Line 138:
# return True, generated_text, create_prompt(passcode)

# In return_prompt Line 142:
# return False, str(e), None
